"""
Servicio para generar reportes e historiales de tareas completadas.
Contiene toda la lógica de negocio para consultas, agregación de datos y generación de archivos.
"""

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, func, desc
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Tuple
from fastapi import HTTPException
import pandas as pd
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch, cm

from app.models.tarea import Tarea
from app.models.usuario import Usuario
from app.models.detalle_tarea import DetalleTarea
from app.models.producto import Producto
from app.models.punto_reposicion import PuntoReposicion
from app.models.estado_tarea import EstadoTarea
from app.models.supervision import Supervision
from app.models.usuario import RolEnum


class ReportesService:
    """Servicio para generar reportes de tareas completadas."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def obtener_historial_tareas_reponedor(
        self,
        id_reponedor: int,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        estado_filtro: Optional[str] = None,
        limit: int = 100,
        offset: int = 0
    ) -> Dict[str, Any]:
        """
        Obtiene el historial de tareas de un reponedor específico con filtros.
        
        Args:
            id_reponedor: ID del reponedor
            fecha_inicio: Fecha de inicio del filtro (opcional)
            fecha_fin: Fecha de fin del filtro (opcional)
            estado_filtro: Estado de tarea para filtrar (opcional)
            limit: Límite de resultados
            offset: Offset para paginación
            
        Returns:
            Dict con el historial de tareas y metadatos
        """
        # Validar que el reponedor existe
        reponedor = self.db.query(Usuario).filter(
            Usuario.id_usuario == id_reponedor,
            Usuario.rol.has(nombre_rol=RolEnum.REPONEDOR.value)
        ).first()
        
        if not reponedor:
            raise HTTPException(
                status_code=404, 
                detail="Reponedor no encontrado o no tiene el rol correcto."
            )
        
        # Construir query base
        query = self.db.query(Tarea).options(
            joinedload(Tarea.detalles)
        ).filter(Tarea.id_reponedor == id_reponedor)
        
        # Aplicar filtros
        if fecha_inicio:
            query = query.filter(Tarea.fecha_creacion >= fecha_inicio)
        
        if fecha_fin:
            query = query.filter(Tarea.fecha_creacion <= fecha_fin)
        
        if estado_filtro:
            estado = self.db.query(EstadoTarea).filter(
                EstadoTarea.nombre_estado.ilike(estado_filtro)
            ).first()
            if estado:
                query = query.filter(Tarea.estado_id == estado.estado_id)
        
        # Obtener total de registros para paginación
        total_registros = query.count()
        
        # Aplicar paginación y ordenamiento
        tareas = query.order_by(desc(Tarea.fecha_creacion)).offset(offset).limit(limit).all()
        
        # Procesar resultados
        historial = []
        estadisticas = {
            "total_tareas": total_registros,
            "tareas_completadas": 0,
            "tareas_canceladas": 0,
            "tareas_en_progreso": 0,
            "total_productos_procesados": 0,
            "promedio_productos_por_tarea": 0
        }
        
        for tarea in tareas:
            # Procesar detalles de la tarea
            productos = []
            total_cantidad = 0
            
            for detalle in tarea.detalles:
                # Obtener producto usando el ID
                producto = self.db.query(Producto).filter(
                    Producto.id_producto == detalle.id_producto
                ).first()
                
                # Obtener punto de reposición usando el ID
                punto_reposicion = self.db.query(PuntoReposicion).filter(
                    PuntoReposicion.id_punto == detalle.id_punto
                ).first()
                
                producto_info = {
                    "id_producto": detalle.id_producto,
                    "nombre_producto": producto.nombre if producto else None,
                    "cantidad": detalle.cantidad,
                    "ubicacion": {
                        "id_punto": detalle.id_punto,
                        "estanteria": punto_reposicion.estanteria if punto_reposicion else None,
                        "nivel": punto_reposicion.nivel if punto_reposicion else None
                    }
                }
                productos.append(producto_info)
                total_cantidad += detalle.cantidad
            
            # Información de la tarea
            tarea_info = {
                "id_tarea": tarea.id_tarea,
                "fecha_creacion": tarea.fecha_creacion.isoformat(),
                "fecha_completada": tarea.fecha_hora_completada.isoformat() if tarea.fecha_hora_completada else None,
                "estado": self._obtener_nombre_estado(tarea.estado_id),
                "supervisor": self._obtener_nombre_supervisor(tarea.id_supervisor),
                "productos": productos,
                "total_productos": len(productos),
                "total_cantidad": total_cantidad,
                "tiempo_estimado_minutos": self._calcular_tiempo_estimado(tarea)
            }
            
            historial.append(tarea_info)
            
            # Actualizar estadísticas
            estado_nombre = self._obtener_nombre_estado(tarea.estado_id)
            if estado_nombre == "completada":
                estadisticas["tareas_completadas"] += 1
            elif estado_nombre == "cancelada":
                estadisticas["tareas_canceladas"] += 1
            elif estado_nombre in ["pendiente", "en progreso"]:
                estadisticas["tareas_en_progreso"] += 1
            
            estadisticas["total_productos_procesados"] += total_cantidad
        
        # Calcular promedio
        if total_registros > 0:
            estadisticas["promedio_productos_por_tarea"] = round(
                sum(len(t["productos"]) for t in historial) / len(historial), 2
            ) if historial else 0
        
        return {
            "reponedor": {
                "id": reponedor.id_usuario,
                "nombre": reponedor.nombre,
                "email": reponedor.correo
            },
            "filtros_aplicados": {
                "fecha_inicio": fecha_inicio.isoformat() if fecha_inicio else None,
                "fecha_fin": fecha_fin.isoformat() if fecha_fin else None,
                "estado": estado_filtro
            },
            "paginacion": {
                "total_registros": total_registros,
                "limite": limit,
                "offset": offset,
                "pagina_actual": (offset // limit) + 1,
                "total_paginas": (total_registros + limit - 1) // limit
            },
            "estadisticas": estadisticas,
            "historial": historial
        }
    
    def generar_reporte_excel(
        self,
        id_reponedor: int,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        estado_filtro: Optional[str] = None
    ) -> Tuple[io.BytesIO, str]:
        """
        Genera un reporte en formato Excel (.xlsx) del historial de tareas.
        
        Returns:
            Tuple con el archivo BytesIO y el nombre del archivo
        """
        # Obtener datos sin paginación para el reporte completo
        datos = self.obtener_historial_tareas_reponedor(
            id_reponedor=id_reponedor,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado_filtro=estado_filtro,
            limit=10000,  # Límite alto para obtener todos los datos
            offset=0
        )
        
        # Crear DataFrame principal
        filas_principales = []
        filas_detalle = []
        
        for tarea in datos["historial"]:
            # Fila principal de la tarea
            fila_principal = {
                "ID Tarea": tarea["id_tarea"],
                "Fecha Creación": tarea["fecha_creacion"][:10],  # Solo la fecha
                "Fecha Completada": tarea["fecha_completada"][:10] if tarea["fecha_completada"] else "",
                "Estado": tarea["estado"],
                "Supervisor": tarea["supervisor"],
                "Total Productos": tarea["total_productos"],
                "Total Cantidad": tarea["total_cantidad"],
                "Tiempo Estimado (min)": tarea["tiempo_estimado_minutos"]
            }
            filas_principales.append(fila_principal)
            
            # Filas de detalle por producto
            for producto in tarea["productos"]:
                fila_detalle = {
                    "ID Tarea": tarea["id_tarea"],
                    "Producto": producto["nombre_producto"],
                    "Cantidad": producto["cantidad"],
                    "Estantería": producto["ubicacion"]["estanteria"],
                    "Nivel": producto["ubicacion"]["nivel"],
                    "ID Punto": producto["ubicacion"]["id_punto"]
                }
                filas_detalle.append(fila_detalle)
        
        # Crear archivo Excel en memoria
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Hoja de resumen
            df_resumen = pd.DataFrame([datos["estadisticas"]])
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Hoja de tareas
            df_tareas = pd.DataFrame(filas_principales)
            df_tareas.to_excel(writer, sheet_name='Tareas', index=False)
            
            # Hoja de productos detallados
            df_productos = pd.DataFrame(filas_detalle)
            df_productos.to_excel(writer, sheet_name='Productos Detalle', index=False)
            
            # Hoja de información del reponedor
            info_reponedor = pd.DataFrame([{
                "Campo": k.replace("_", " ").title(),
                "Valor": v
            } for k, v in datos["reponedor"].items()])
            info_reponedor.to_excel(writer, sheet_name='Información Reponedor', index=False)
            
            # Aplicar estilos mejorados
            self._aplicar_estilos_excel_avanzados(writer, datos)
        
        buffer.seek(0)
        
        # Generar nombre del archivo
        reponedor_nombre = datos["reponedor"]["nombre"].replace(" ", "_")
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"historial_{reponedor_nombre}_{fecha_actual}.xlsx"
        
        return buffer, nombre_archivo
    
    def generar_reporte_pdf(
        self,
        id_reponedor: int,
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        estado_filtro: Optional[str] = None
    ) -> Tuple[io.BytesIO, str]:
        """
        Genera un reporte en formato PDF del historial de tareas.
        
        Returns:
            Tuple con el archivo BytesIO y el nombre del archivo
        """
        # Obtener datos
        datos = self.obtener_historial_tareas_reponedor(
            id_reponedor=id_reponedor,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado_filtro=estado_filtro,
            limit=10000,
            offset=0
        )
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
        story = []
        
        # Estilos mejorados
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Centrado
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=20,
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold',
            borderWidth=2,
            borderColor=colors.HexColor('#d5e3f0'),
            borderPadding=5,
            backColor=colors.HexColor('#f8f9fa')
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica'
        )
        
        # === ENCABEZADO DEL DOCUMENTO ===
        # Logo/Título principal
        story.append(Paragraph("POE - PATH OPTIMIZATION ENGINE", title_style))
        story.append(Paragraph("Reporte de Rendimiento de Reponedor", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Línea separadora
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1f4e79')))
        story.append(Spacer(1, 20))
        
        # === INFORMACIÓN DEL REPONEDOR ===
        story.append(Paragraph("📋 INFORMACIÓN DEL REPONEDOR", section_style))
        
        reponedor_data = [
            ["Campo", "Información"],
            ["👤 Nombre Completo", datos["reponedor"]["nombre"]],
            ["📧 Correo Electrónico", datos["reponedor"]["email"]],
            ["🆔 ID de Usuario", str(datos["reponedor"]["id"])],
            ["📅 Fecha del Reporte", datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]
        
        reponedor_table = Table(reponedor_data, colWidths=[4*cm, 12*cm])
        reponedor_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#d5e3f0')),
            ('BACKGROUND', (1, 1), (1, -1), colors.white),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1f4e79')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(reponedor_table)
        story.append(Spacer(1, 25))
        
        # === ESTADÍSTICAS GENERALES ===
        story.append(Paragraph("📊 RESUMEN ESTADÍSTICO", section_style))
        
        stats = datos["estadisticas"]
        stats_data = [
            ["Métrica", "Valor", "Indicador"],
            ["📝 Total de Tareas", str(stats["total_tareas"]), ""],
            ["✅ Tareas Completadas", str(stats["tareas_completadas"]), self._get_performance_indicator(stats["tareas_completadas"], stats["total_tareas"])],
            ["❌ Tareas Canceladas", str(stats["tareas_canceladas"]), ""],
            ["🔄 Tareas en Progreso", str(stats["tareas_en_progreso"]), ""],
            ["📦 Total Productos", str(stats["total_productos_procesados"]), ""],
            ["📈 Promedio Productos/Tarea", f"{stats['promedio_productos_por_tarea']:.1f}", ""]
        ]
        
        stats_table = Table(stats_data, colWidths=[6*cm, 4*cm, 6*cm])
        stats_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70ad47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f8f0')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#70ad47')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            
            # Colores especiales para filas
            ('BACKGROUND', (1, 2), (1, 2), colors.HexColor('#d4edda')),  # Completadas
            ('BACKGROUND', (1, 3), (1, 3), colors.HexColor('#f8d7da')),  # Canceladas
            ('BACKGROUND', (1, 4), (1, 4), colors.HexColor('#fff3cd')),  # En progreso
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 25))
        
        # === HISTORIAL DE TAREAS ===
        if datos["historial"]:
            story.append(Paragraph("📋 HISTORIAL DE TAREAS", section_style))
            
            # Nota sobre limitación
            total_tareas = len(datos["historial"])
            tareas_mostrar = min(30, total_tareas)
            
            if total_tareas > 30:
                story.append(Paragraph(
                    f"<b>Nota:</b> Se muestran las {tareas_mostrar} tareas más recientes de {total_tareas} totales.",
                    normal_style
                ))
                story.append(Spacer(1, 10))
            
            # Preparar datos de la tabla
            task_data = [["ID", "Fecha Creación", "Estado", "Productos", "Cantidad", "Supervisor"]]
            
            for tarea in datos["historial"][:tareas_mostrar]:
                # Formatear estado con emoji
                estado = tarea["estado"]
                if "completada" in estado.lower():
                    estado_formatted = f"✅ {estado}"
                elif "cancelada" in estado.lower():
                    estado_formatted = f"❌ {estado}"
                elif "progreso" in estado.lower():
                    estado_formatted = f"🔄 {estado}"
                else:
                    estado_formatted = f"📝 {estado}"
                
                task_data.append([
                    str(tarea["id_tarea"]),
                    tarea["fecha_creacion"][:10],
                    estado_formatted,
                    str(tarea["total_productos"]),
                    str(tarea["total_cantidad"]),
                    tarea["supervisor"][:15] + "..." if len(tarea["supervisor"]) > 15 else tarea["supervisor"]
                ])
            
            # Ajustar ancho de columnas para que quepan en la página
            task_table = Table(task_data, colWidths=[2*cm, 2.5*cm, 3.5*cm, 2*cm, 2*cm, 4*cm])
            task_table.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f79646')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#f79646')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Alternar colores de filas
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#fff8f0')),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fff8f0')),
                ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#fff8f0')),
            ]))
            
            # Aplicar colores según estado a las filas de datos
            for i, tarea in enumerate(datos["historial"][:tareas_mostrar], 1):
                estado = tarea["estado"].lower()
                if "completada" in estado:
                    task_table.setStyle(TableStyle([
                        ('BACKGROUND', (2, i), (2, i), colors.HexColor('#d4edda'))
                    ]))
                elif "cancelada" in estado:
                    task_table.setStyle(TableStyle([
                        ('BACKGROUND', (2, i), (2, i), colors.HexColor('#f8d7da'))
                    ]))
                elif "progreso" in estado:
                    task_table.setStyle(TableStyle([
                        ('BACKGROUND', (2, i), (2, i), colors.HexColor('#fff3cd'))
                    ]))
            
            story.append(task_table)
        
        story.append(Spacer(1, 30))
        
        # === PIE DE PÁGINA ===
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.grey
        )
        
        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Reporte generado automáticamente por POE - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            footer_style
        ))
        story.append(Paragraph("© 2025 Path Optimization Engine - Sistema de Gestión de Inventario", footer_style))
        
        # Construir el PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generar nombre del archivo
        reponedor_nombre = datos["reponedor"]["nombre"].replace(" ", "_")
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"historial_{reponedor_nombre}_{fecha_actual}.pdf"
        
        return buffer, nombre_archivo
    
    def validar_acceso_supervisor(self, supervisor_id: int, reponedor_id: int) -> bool:
        """
        Valida si un supervisor tiene acceso a los datos de un reponedor específico.
        
        Args:
            supervisor_id: ID del supervisor
            reponedor_id: ID del reponedor
            
        Returns:
            True si tiene acceso, False en caso contrario
        """
        supervision = self.db.query(Supervision).filter(
            Supervision.supervisor_id == supervisor_id,
            Supervision.reponedor_id == reponedor_id
        ).first()
        
        return supervision is not None
    
    def obtener_reponedores_supervisor(self, supervisor_id: int) -> List[Dict[str, Any]]:
        """
        Obtiene la lista de reponedores bajo la supervisión de un supervisor.
        
        Args:
            supervisor_id: ID del supervisor
            
        Returns:
            Lista de reponedores con información básica
        """
        supervisiones = self.db.query(Supervision).options(
            joinedload(Supervision.reponedor)
        ).filter(Supervision.supervisor_id == supervisor_id).all()
        
        reponedores = []
        for supervision in supervisiones:
            if supervision.reponedor:
                reponedores.append({
                    "id_usuario": supervision.reponedor.id_usuario,
                    "nombre": supervision.reponedor.nombre,
                    "email": supervision.reponedor.correo,
                    "estado": supervision.reponedor.estado
                })
        
        return reponedores
    
    def obtener_productos_mas_repuestos(
        self,
        fecha_inicio: date,
        fecha_fin: date,
        limit: int = 100
    ) -> Dict[str, Any]:
        """
        Obtiene los productos más repuestos en un rango de fechas.
        
        Args:
            fecha_inicio: Fecha de inicio del filtro
            fecha_fin: Fecha de fin del filtro
            limit: Límite de productos a retornar
            
        Returns:
            Dict con los productos más repuestos y estadísticas
        """
        # Validar fechas
        if fecha_inicio > fecha_fin:
            raise HTTPException(
                status_code=400,
                detail="La fecha de inicio no puede ser mayor que la fecha de fin."
            )
        
        # Query para obtener productos más repuestos
        productos_query = (
            self.db.query(
                Producto.id_producto,
                Producto.nombre,
                Producto.categoria,
                func.sum(DetalleTarea.cantidad).label('cantidad_total_repuesta'),
                func.count(DetalleTarea.id_detalle).label('numero_reposiciones')
            )
            .join(DetalleTarea, DetalleTarea.id_producto == Producto.id_producto)
            .join(Tarea, Tarea.id_tarea == DetalleTarea.id_tarea)
            .filter(
                and_(
                    Tarea.fecha_creacion >= fecha_inicio,
                    Tarea.fecha_creacion <= fecha_fin
                )
            )
            .group_by(Producto.id_producto, Producto.nombre, Producto.categoria)
            .order_by(func.sum(DetalleTarea.cantidad).desc())
            .limit(limit)
        )
        
        productos_data = productos_query.all()
        
        # Procesar resultados
        productos = []
        total_productos_repuestos = 0
        total_cantidad_general = 0
        
        for producto in productos_data:
            producto_info = {
                "id_producto": producto.id_producto,
                "nombre": producto.nombre,
                "categoria": producto.categoria,
                "cantidad_total_repuesta": int(producto.cantidad_total_repuesta),
                "numero_reposiciones": int(producto.numero_reposiciones)
            }
            productos.append(producto_info)
            total_productos_repuestos += 1
            total_cantidad_general += int(producto.cantidad_total_repuesta)
        
        # Estadísticas adicionales
        estadisticas = {
            "total_productos_repuestos": total_productos_repuestos,
            "cantidad_total_general": total_cantidad_general,
            "promedio_cantidad_por_producto": round(
                total_cantidad_general / total_productos_repuestos, 2
            ) if total_productos_repuestos > 0 else 0,
            "promedio_reposiciones_por_producto": round(
                sum(p["numero_reposiciones"] for p in productos) / total_productos_repuestos, 2
            ) if total_productos_repuestos > 0 else 0
        }
        
        # Análisis por categorías
        categorias_stats = {}
        for producto in productos:
            categoria = producto["categoria"]
            if categoria not in categorias_stats:
                categorias_stats[categoria] = {
                    "productos_count": 0,
                    "cantidad_total": 0,
                    "reposiciones_total": 0
                }
            
            categorias_stats[categoria]["productos_count"] += 1
            categorias_stats[categoria]["cantidad_total"] += producto["cantidad_total_repuesta"]
            categorias_stats[categoria]["reposiciones_total"] += producto["numero_reposiciones"]
        
        return {
            "filtros": {
                "fecha_inicio": fecha_inicio.isoformat(),
                "fecha_fin": fecha_fin.isoformat(),
                "limite_productos": limit
            },
            "estadisticas": estadisticas,
            "analisis_por_categorias": categorias_stats,
            "productos": productos,
            "metadatos": {
                "generado_el": datetime.now().isoformat(),
                "total_registros": len(productos)
            }
        }

    def generar_reporte_productos_excel(
        self,
        fecha_inicio: date,
        fecha_fin: date,
        limit: int = 100
    ) -> Tuple[io.BytesIO, str]:
        """
        Genera un reporte en formato Excel de productos más repuestos.
        
        Returns:
            Tuple con el archivo BytesIO y el nombre del archivo
        """
        # Obtener datos
        datos = self.obtener_productos_mas_repuestos(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            limit=limit
        )
        
        # Crear archivo Excel en memoria
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Hoja principal - Productos más repuestos
            df_productos = pd.DataFrame(datos["productos"])
            df_productos.to_excel(writer, sheet_name='Productos Más Repuestos', index=False)
            
            # Hoja de estadísticas generales
            stats_data = []
            for key, value in datos["estadisticas"].items():
                stats_data.append({
                    "Métrica": key.replace("_", " ").title(),
                    "Valor": value
                })
            df_stats = pd.DataFrame(stats_data)
            df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            # Hoja de análisis por categorías
            categorias_data = []
            for categoria, stats in datos["analisis_por_categorias"].items():
                categorias_data.append({
                    "Categoría": categoria,
                    "Productos": stats["productos_count"],
                    "Cantidad Total": stats["cantidad_total"],
                    "Reposiciones Total": stats["reposiciones_total"],
                    "Promedio Cantidad/Producto": round(stats["cantidad_total"] / stats["productos_count"], 2) if stats["productos_count"] > 0 else 0
                })
            df_categorias = pd.DataFrame(categorias_data)
            df_categorias.to_excel(writer, sheet_name='Análisis por Categorías', index=False)
            
            # Hoja de metadatos del reporte
            metadata = []
            metadata.append({"Campo": "Fecha Inicio", "Valor": datos["filtros"]["fecha_inicio"]})
            metadata.append({"Campo": "Fecha Fin", "Valor": datos["filtros"]["fecha_fin"]})
            metadata.append({"Campo": "Límite Productos", "Valor": datos["filtros"]["limite_productos"]})
            metadata.append({"Campo": "Fecha Generación", "Valor": datos["metadatos"]["generado_el"]})
            metadata.append({"Campo": "Total Registros", "Valor": datos["metadatos"]["total_registros"]})
            
            df_metadata = pd.DataFrame(metadata)
            df_metadata.to_excel(writer, sheet_name='Información del Reporte', index=False)
            
            # Aplicar estilos mejorados
            self._aplicar_estilos_productos_excel(writer, datos)
        
        buffer.seek(0)
        
        # Generar nombre del archivo
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        fecha_inicio_str = fecha_inicio.strftime("%Y%m%d")
        fecha_fin_str = fecha_fin.strftime("%Y%m%d")
        nombre_archivo = f"productos_mas_repuestos_{fecha_inicio_str}_{fecha_fin_str}_{fecha_actual}.xlsx"
        
        return buffer, nombre_archivo

    def generar_reporte_productos_pdf(
        self,
        fecha_inicio: date,
        fecha_fin: date,
        limit: int = 100
    ) -> Tuple[io.BytesIO, str]:
        """
        Genera un reporte en formato PDF de productos más repuestos.
        
        Returns:
            Tuple con el archivo BytesIO y el nombre del archivo
        """
        # Obtener datos
        datos = self.obtener_productos_mas_repuestos(
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            limit=limit
        )
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Centrado
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold'
        )
        
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=15,
            spaceBefore=20,
            textColor=colors.HexColor('#1f4e79'),
            fontName='Helvetica-Bold',
            borderWidth=2,
            borderColor=colors.HexColor('#d5e3f0'),
            borderPadding=5,
            backColor=colors.HexColor('#f8f9fa')
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica'
        )
        
        # === ENCABEZADO DEL DOCUMENTO ===
        story.append(Paragraph("POE - PATH OPTIMIZATION ENGINE", title_style))
        story.append(Paragraph("Reporte de Productos Más Repuestos", subtitle_style))
        story.append(Spacer(1, 20))
        
        # Línea separadora
        from reportlab.platypus import HRFlowable
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1f4e79')))
        story.append(Spacer(1, 20))
        
        # === INFORMACIÓN DEL REPORTE ===
        story.append(Paragraph("📋 INFORMACIÓN DEL REPORTE", section_style))
        
        info_data = [
            ["Parámetro", "Valor"],
            ["📅 Fecha de Inicio", datos["filtros"]["fecha_inicio"]],
            ["📅 Fecha de Fin", datos["filtros"]["fecha_fin"]],
            ["🔢 Límite de Productos", str(datos["filtros"]["limite_productos"])],
            ["📊 Total de Registros", str(datos["metadatos"]["total_registros"])],
            ["⏰ Fecha de Generación", datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 12*cm])
        info_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#d5e3f0')),
            ('BACKGROUND', (1, 1), (1, -1), colors.white),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#1f4e79')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 25))
        
        # === ESTADÍSTICAS GENERALES ===
        story.append(Paragraph("📊 ESTADÍSTICAS GENERALES", section_style))
        
        stats = datos["estadisticas"]
        stats_data = [
            ["Métrica", "Valor"],
            ["📦 Total Productos Repuestos", str(stats["total_productos_repuestos"])],
            ["📈 Cantidad Total General", str(stats["cantidad_total_general"])],
            ["📊 Promedio Cantidad/Producto", f"{stats['promedio_cantidad_por_producto']:.2f}"],
            ["🔄 Promedio Reposiciones/Producto", f"{stats['promedio_reposiciones_por_producto']:.2f}"]
        ]
        
        stats_table = Table(stats_data, colWidths=[8*cm, 8*cm])
        stats_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70ad47')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0f8f0')),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#70ad47')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 25))
        
        # === TOP PRODUCTOS MÁS REPUESTOS ===
        if datos["productos"]:
            story.append(Paragraph("🏆 TOP PRODUCTOS MÁS REPUESTOS", section_style))
            
            # Limitar a los primeros 20 productos para que quepan en PDF
            productos_mostrar = datos["productos"][:20]
            
            if len(datos["productos"]) > 20:
                story.append(Paragraph(
                    f"<b>Nota:</b> Se muestran los top {len(productos_mostrar)} productos de {len(datos['productos'])} totales.",
                    normal_style
                ))
                story.append(Spacer(1, 10))
            
            # Preparar datos de la tabla
            productos_data = [["Rank", "Producto", "Categoría", "Cantidad", "Reposiciones"]]
            
            for i, producto in enumerate(productos_mostrar, 1):
                productos_data.append([
                    str(i),
                    producto["nombre"][:25] + "..." if len(producto["nombre"]) > 25 else producto["nombre"],
                    producto["categoria"],
                    str(producto["cantidad_total_repuesta"]),
                    str(producto["numero_reposiciones"])
                ])
            
            productos_table = Table(productos_data, colWidths=[1.5*cm, 6*cm, 3*cm, 2.5*cm, 3*cm])
            productos_table.setStyle(TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f79646')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#f79646')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Alternar colores de filas
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#fff8f0')),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fff8f0')),
                ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#fff8f0')),
            ]))
            
            # Resaltar top 3
            for i in range(1, min(4, len(productos_mostrar) + 1)):
                productos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ffd700' if i == 1 else '#c0c0c0' if i == 2 else '#cd7f32')),
                    ('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold')
                ]))
            
            story.append(productos_table)
        
        story.append(Spacer(1, 30))
        
        # === PIE DE PÁGINA ===
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.grey
        )
        
        story.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Reporte generado automáticamente por POE - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            footer_style
        ))
        story.append(Paragraph("© 2025 Path Optimization Engine - Sistema de Gestión de Inventario", footer_style))
        
        # Construir el PDF
        doc.build(story)
        buffer.seek(0)
        
        # Generar nombre del archivo
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        fecha_inicio_str = fecha_inicio.strftime("%Y%m%d")
        fecha_fin_str = fecha_fin.strftime("%Y%m%d")
        nombre_archivo = f"productos_mas_repuestos_{fecha_inicio_str}_{fecha_fin_str}_{fecha_actual}.pdf"
        
        return buffer, nombre_archivo

    def _obtener_nombre_estado(self, estado_id: int) -> Optional[str]:
        """
        Obtiene el nombre del estado por su ID.
        
        Args:
            estado_id: ID del estado
            
        Returns:
            Nombre del estado o None si no se encuentra
        """
        estado = self.db.query(EstadoTarea).filter(
            EstadoTarea.estado_id == estado_id
        ).first()
        return estado.nombre_estado if estado else None
    
    def _obtener_nombre_supervisor(self, supervisor_id: int) -> Optional[str]:
        """
        Obtiene el nombre del supervisor por su ID.
        
        Args:
            supervisor_id: ID del supervisor
            
        Returns:
            Nombre del supervisor o None si no se encuentra
        """
        supervisor = self.db.query(Usuario).filter(
            Usuario.id_usuario == supervisor_id
        ).first()
        return supervisor.nombre if supervisor else None
    
    def _calcular_tiempo_estimado(self, tarea: Tarea) -> Optional[float]:
        """
        Calcula el tiempo estimado para completar una tarea.
        
        Args:
            tarea: Objeto Tarea
            
        Returns:
            Tiempo estimado en minutos o None si no se puede calcular
        """
        if not tarea.fecha_hora_completada or not tarea.fecha_creacion:
            return None
        
        # Convertir fecha_creacion a datetime para la comparación
        if isinstance(tarea.fecha_creacion, date):
            fecha_inicio = datetime.combine(tarea.fecha_creacion, datetime.min.time())
        else:
            fecha_inicio = tarea.fecha_creacion
        
        diferencia = tarea.fecha_hora_completada - fecha_inicio
        return round(diferencia.total_seconds() / 60, 2)
    
    def _aplicar_estilos_excel_avanzados(self, writer, datos):
        """Aplica estilos avanzados y profesionales al archivo Excel."""
        from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.chart import BarChart, Reference
        
        # Colores corporativos
        azul_principal = "1f4e79"
        azul_claro = "d5e3f0"
        verde_exito = "70ad47"
        naranja_advertencia = "f79646"
        rojo_error = "c55a5a"
        gris_claro = "f2f2f2"
        
        # Estilos base
        font_titulo = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        font_subtitulo = Font(name='Calibri', size=12, bold=True, color=azul_principal)
        font_encabezado = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        font_normal = Font(name='Calibri', size=10)
        font_resaltado = Font(name='Calibri', size=10, bold=True)
        
        fill_titulo = PatternFill(start_color=azul_principal, end_color=azul_principal, fill_type="solid")
        fill_encabezado = PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type="solid")
        fill_datos_par = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_datos_impar = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
        
        border_fino = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        border_grueso = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        alignment_centro = Alignment(horizontal='center', vertical='center')
        alignment_izquierda = Alignment(horizontal='left', vertical='center')
        
        # === HOJA DE RESUMEN ===
        ws_resumen = writer.sheets['Resumen']
        
        # Título principal
        ws_resumen.insert_rows(1, 3)
        ws_resumen['A1'] = "REPORTE DE RENDIMIENTO - POE"
        ws_resumen['A1'].font = Font(name='Calibri', size=18, bold=True, color=azul_principal)
        ws_resumen['A1'].alignment = alignment_centro
        ws_resumen.merge_cells('A1:F1')
        
        ws_resumen['A2'] = f"Reponedor: {datos['reponedor']['nombre']}"
        ws_resumen['A2'].font = font_subtitulo
        ws_resumen['A2'].alignment = alignment_centro
        ws_resumen.merge_cells('A2:F2')
        
        ws_resumen['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_resumen['A3'].font = font_normal
        ws_resumen['A3'].alignment = alignment_centro
        ws_resumen.merge_cells('A3:F3')
        
        # Formatear estadísticas
        for row in range(4, ws_resumen.max_row + 1):
            for col in range(1, ws_resumen.max_column + 1):
                cell = ws_resumen.cell(row=row, column=col)
                if row == 4:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    cell.alignment = alignment_centro if col > 1 else alignment_izquierda
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_resumen.column_dimensions['A'].width = 25
        for col in range(2, ws_resumen.max_column + 1):
            ws_resumen.column_dimensions[chr(64 + col)].width = 15
        
        # === HOJA DE TAREAS ===
        ws_tareas = writer.sheets['Tareas']
        
        # Título
        ws_tareas.insert_rows(1, 2)
        ws_tareas['A1'] = "HISTORIAL DE TAREAS"
        ws_tareas['A1'].font = font_titulo
        ws_tareas['A1'].fill = fill_titulo
        ws_tareas['A1'].alignment = alignment_centro
        ws_tareas.merge_cells(f'A1:{chr(64 + ws_tareas.max_column)}1')
        
        # Formatear tabla de tareas
        for row in range(3, ws_tareas.max_row + 1):
            for col in range(1, ws_tareas.max_column + 1):
                cell = ws_tareas.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.alignment = alignment_centro
                    cell.border = border_fino
                    
                    # Colorear según estado
                    if col == 4:  # Columna de estado
                        estado = str(cell.value).lower()
                        if 'completada' in estado:
                            cell.fill = PatternFill(start_color=verde_exito, end_color=verde_exito, fill_type="solid")
                            cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
                        elif 'cancelada' in estado:
                            cell.fill = PatternFill(start_color=rojo_error, end_color=rojo_error, fill_type="solid")
                            cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
                        elif 'progreso' in estado or 'pendiente' in estado:
                            cell.fill = PatternFill(start_color=naranja_advertencia, end_color=naranja_advertencia, fill_type="solid")
                            cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
                        else:
                            cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    else:
                        cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
        
        # Ajustar ancho de columnas
        column_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 15, 'E': 20, 'F': 15, 'G': 15, 'H': 18}
        for col, width in column_widths.items():
            ws_tareas.column_dimensions[col].width = width
        
        # === HOJA DE PRODUCTOS DETALLE ===
        ws_productos = writer.sheets['Productos Detalle']
        
        # Título
        ws_productos.insert_rows(1, 2)
        ws_productos['A1'] = "DETALLE DE PRODUCTOS POR TAREA"
        ws_productos['A1'].font = font_titulo
        ws_productos['A1'].fill = fill_titulo
        ws_productos['A1'].alignment = alignment_centro
        ws_productos.merge_cells(f'A1:{chr(64 + ws_productos.max_column)}1')
        
        # Formatear tabla de productos
        for row in range(3, ws_productos.max_row + 1):
            for col in range(1, ws_productos.max_column + 1):
                cell = ws_productos.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.alignment = alignment_centro if col in [1, 3, 6] else alignment_izquierda
                    cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_productos.column_dimensions['A'].width = 12
        ws_productos.column_dimensions['B'].width = 30
        ws_productos.column_dimensions['C'].width = 12
        ws_productos.column_dimensions['D'].width = 15
        ws_productos.column_dimensions['E'].width = 12
        ws_productos.column_dimensions['F'].width = 12
        
        # === HOJA DE INFORMACIÓN DEL REPONEDOR ===
        ws_info = writer.sheets['Información Reponedor']
        
        # Título
        ws_info.insert_rows(1, 2)
        ws_info['A1'] = "INFORMACIÓN DEL REPONEDOR"
        ws_info['A1'].font = font_titulo
        ws_info['A1'].fill = fill_titulo
        ws_info['A1'].alignment = alignment_centro
        ws_info.merge_cells('A1:B1')
        
        # Formatear información
        for row in range(3, ws_info.max_row + 1):
            for col in range(1, ws_info.max_column + 1):
                cell = ws_info.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    if col == 1:
                        cell.font = font_resaltado
                        cell.fill = fill_encabezado
                        cell.alignment = alignment_izquierda
                    else:
                        cell.font = font_normal
                        cell.fill = fill_datos_par
                        cell.alignment = alignment_izquierda
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 30
    
    def _get_performance_indicator(self, completadas: int, total: int) -> str:
        """Genera un indicador visual de rendimiento."""
        if total == 0:
            return "N/A"
        
        porcentaje = (completadas / total) * 100
        
        if porcentaje >= 90:
            return "🟢 Excelente"
        elif porcentaje >= 75:
            return "🟡 Bueno"
        elif porcentaje >= 50:
            return "🟠 Regular"
        else:
            return "🔴 Necesita Mejorar"
    
    def _aplicar_estilos_productos_excel(self, writer, datos):
        """Aplica estilos específicos para el reporte de productos más repuestos."""
        from openpyxl.styles import Font, Fill, Border, Side, Alignment, PatternFill
        from openpyxl.chart import BarChart, Reference
        
        # Colores corporativos
        azul_principal = "1f4e79"
        azul_claro = "d5e3f0"
        verde_exito = "70ad47"
        naranja_advertencia = "f79646"
        rojo_error = "c55a5a"
        gris_claro = "f2f2f2"
        
        # Estilos base
        font_titulo = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        font_subtitulo = Font(name='Calibri', size=12, bold=True, color=azul_principal)
        font_encabezado = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        font_normal = Font(name='Calibri', size=10)
        font_resaltado = Font(name='Calibri', size=10, bold=True)
        
        fill_titulo = PatternFill(start_color=azul_principal, end_color=azul_principal, fill_type="solid")
        fill_encabezado = PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type="solid")
        fill_datos_par = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_datos_impar = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
        
        border_fino = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        border_grueso = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        alignment_centro = Alignment(horizontal='center', vertical='center')
        alignment_izquierda = Alignment(horizontal='left', vertical='center')
        
        # === HOJA DE PRODUCTOS ===
        ws_productos = writer.sheets['Productos Más Repuestos']
        
        # Título principal
        ws_productos.insert_rows(1, 3)
        ws_productos['A1'] = "REPORTE DE PRODUCTOS MÁS REPUESTOS - POE"
        ws_productos['A1'].font = Font(name='Calibri', size=18, bold=True, color=azul_principal)
        ws_productos['A1'].alignment = alignment_centro
        ws_productos.merge_cells('A1:E1')
        
        ws_productos['A2'] = f"Período: {datos['filtros']['fecha_inicio']} al {datos['filtros']['fecha_fin']}"
        ws_productos['A2'].font = font_subtitulo
        ws_productos['A2'].alignment = alignment_centro
        ws_productos.merge_cells('A2:E2')
        
        ws_productos['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_productos['A3'].font = font_normal
        ws_productos['A3'].alignment = alignment_centro
        ws_productos.merge_cells('A3:E3')
        
        # Formatear tabla de productos
        for row in range(4, ws_productos.max_row + 1):
            for col in range(1, ws_productos.max_column + 1):
                cell = ws_productos.cell(row=row, column=col)
                if row == 4:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.alignment = alignment_centro if col in [1, 4, 5] else alignment_izquierda
                    cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_productos.column_dimensions['A'].width = 12
        ws_productos.column_dimensions['B'].width = 40
        ws_productos.column_dimensions['C'].width = 15
        ws_productos.column_dimensions['D'].width = 20
        ws_productos.column_dimensions['E'].width = 18
        
        # === HOJA DE ESTADÍSTICAS ===
        ws_stats = writer.sheets['Estadísticas']
        
        # Título
        ws_stats.insert_rows(1, 2)
        ws_stats['A1'] = "ESTADÍSTICAS GENERALES"
        ws_stats['A1'].font = font_titulo
        ws_stats['A1'].fill = fill_titulo
        ws_stats['A1'].alignment = alignment_centro
        ws_stats.merge_cells('A1:B1')
        
        # Formatear estadísticas
        for row in range(3, ws_stats.max_row + 1):
            for col in range(1, ws_stats.max_column + 1):
                cell = ws_stats.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    if col == 1:
                        cell.font = font_resaltado
                        cell.fill = fill_encabezado
                        cell.alignment = alignment_izquierda
                    else:
                        cell.font = font_normal
                        cell.fill = fill_datos_par
                        cell.alignment = alignment_centro
                    cell.border = border_fino
                cell = ws_stats.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    if col == 1:
                        cell.font = font_resaltado
                        cell.fill = fill_encabezado
                        cell.alignment = alignment_izquierda
                    else:
                        cell.font = font_normal
                        cell.fill = fill_datos_par
                        cell.alignment = alignment_centro
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 20
        
        # === HOJA DE CATEGORÍAS ===
        ws_categorias = writer.sheets['Análisis por Categorías']
        
        # Título
        ws_categorias.insert_rows(1, 2)
        ws_categorias['A1'] = "ANÁLISIS POR CATEGORÍAS"
        ws_categorias['A1'].font = font_titulo
        ws_categorias['A1'].fill = fill_titulo
        ws_categorias['A1'].alignment = alignment_centro
        ws_categorias.merge_cells('A1:D1')
        
        # Formatear tabla de categorías
        for row in range(3, ws_categorias.max_row + 1):
            for col in range(1, ws_categorias.max_column + 1):
                cell = ws_categorias.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.alignment = alignment_centro if col > 1 else alignment_izquierda
                    cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    cell.border = border_fino
            for col in range(1, ws_categorias.max_column + 1):
                cell = ws_categorias.cell(row=row, column=col)
                if row == 3:  # Encabezados
                    cell.font = font_encabezado
                    cell.fill = fill_titulo
                    cell.alignment = alignment_centro
                    cell.border = border_grueso
                else:
                    cell.font = font_normal
                    cell.alignment = alignment_centro if col > 1 else alignment_izquierda
                    cell.fill = fill_datos_par if row % 2 == 0 else fill_datos_impar
                    cell.border = border_fino
        
        # Ajustar ancho de columnas
        ws_categorias.column_dimensions['A'].width = 20
        ws_categorias.column_dimensions['B'].width = 18
        ws_categorias.column_dimensions['C'].width = 18
        ws_categorias.column_dimensions['D'].width = 18
